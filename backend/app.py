# backend/app.py - VERSÃO COM BANCO DE DADOS E AUTENTICAÇÃO

from flask import Flask, request, send_file
from flask_cors import CORS
from io import BytesIO
from datetime import datetime
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image

# NOVAS IMPORTAÇÕES PARA O BANCO DE DADOS E AUTENTICAÇÃO
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_jwt_extended import JWTManager

app = Flask(__name__)
# Sua configuração de CORS para produção
allowed_origins = [
    "https://exportacaohevile.netlify.app" # Seu site de produção
    "http://localhost:3000"               # Seu site de desenvolvimento
]
CORS(app, resources={r"/api/*": {"origins": allowed_origins}})
# ==================================
# --- 1. CONFIGURAÇÃO DO BANCO DE DADOS E AUTENTICAÇÃO ---

# Pega a URL do banco de dados que você configurou no Render
db_url = os.environ.get('DATABASE_URL')
# O Render usa 'postgres://' mas o SQLAlchemy espera 'postgresql://'
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
# IMPORTANTE: Mude esta chave para qualquer frase secreta e complexa
app.config['JWT_SECRET_KEY'] = 'mude-isso-para-uma-chave-secreta-muito-forte' 

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
jwt = JWTManager(app)

# --- 2. DEFINIÇÃO DA TABELA DE USUÁRIOS (CLIENTES) ---
# Aqui definimos quais dados cada cliente terá no banco de dados

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    
    # Dados pré-preenchidos
    shipperName = db.Column(db.String(255), nullable=True)
    shipperInfo = db.Column(db.Text, nullable=True)
    consignee = db.Column(db.Text, nullable=True)
    notifyParty = db.Column(db.Text, nullable=True)
    notifyParty2 = db.Column(db.Text, nullable=True)
    
    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)

# =====================================================================
# O RESTO DO SEU CÓDIGO (Excel, etc.) PERMANECE EXATAMENTE IGUAL
# =====================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@app.route('/api/generate-file', methods=['POST'])
def handle_form():
    if not request.json:
        return {"error": "Missing JSON"}, 400

    data = request.json
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "HBL SAMPLE"

    # Sua configuração de página A4
    sheet.page_setup.orientation = 'portrait'
    # Correção para o atributo de paper_size
    try:
        sheet.page_setup.paper_size = sheet.page_setup.PAPERSIZE_A4
    except AttributeError:
        sheet.page_setup.paper_size = '9' # 9 é o código para A4
        
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    # --- 1. DEFINIÇÕES DE ESTILO ---
    brand_blue_fill = PatternFill(start_color="4c748c", end_color="4c748c", fill_type="solid")
    brand_orange_fill = PatternFill(start_color="f96b06", end_color="f96b06", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=9)
    black_bold_font = Font(color="000000", bold=True, size=9)
    default_font = Font(color="000000", size=9)
    small_font = Font(color="000000", size=8)
    
    thin_black_border = Border(left=Side(style='thin', color='000000'), 
                               right=Side(style='thin', color='000000'), 
                               top=Side(style='thin', color='000000'), 
                               bottom=Side(style='thin', color='000000'))
    
    dotted_border_side = Side(style='dotted', color='000000')
    dotted_border = Border(left=dotted_border_side, right=dotted_border_side, top=dotted_border_side, bottom=dotted_border_side)

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_left_center = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right_bottom = Alignment(horizontal='right', vertical='bottom', wrap_text=True)

    # --- 2. DIMENSÕES (LARGURAS REDUZIDAS PARA CABER EM A4) ---
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 8.5
    sheet.column_dimensions['E'].width = 9
    sheet.column_dimensions['F'].width = 9
    sheet.column_dimensions['G'].width = 9
    sheet.column_dimensions['H'].width = 9
    sheet.column_dimensions['I'].width = 9
    sheet.column_dimensions['J'].width = 9.5
    sheet.column_dimensions['K'].width = 9.5
    sheet.column_dimensions['L'].width = 9.5
    sheet.column_dimensions['M'].width = 9.5
    
    for i in range(1, 90): sheet.row_dimensions[i].height = 13.5
    for i in [2,8,15,31,47]: sheet.row_dimensions[i].height = 16
    for i in range(32, 46): sheet.row_dimensions[i].height = 16

    # --- 3. CONSTRUÇÃO DO TEMPLATE (AJUSTADO PARA COLUNAS A-M) ---
    
    def style_and_merge(cell_range, value, fill, font=black_bold_font, alignment=align_center):
        start_cell_id = cell_range.split(':')[0]
        cell = sheet[start_cell_id]
        cell.value = value
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        sheet.merge_cells(cell_range)

    # -- Seção 1: Cabeçalho Superior --
    style_and_merge('A1:I1', "SHIPPER/EXPORT", brand_orange_fill)
    style_and_merge('J1:K1', "BOOKING Nº", brand_blue_fill, font=white_font)
    style_and_merge('L1:M1', "B/L Nº", brand_blue_fill, font=white_font)
    
    sheet['A2'].value = f"{data.get('shipperName', '')}\n{data.get('shipperInfo', '')}"; sheet['A2'].alignment = align_left_top; sheet.merge_cells('A2:I8')
    sheet['J2'].value = data.get('bookingNo'); sheet['J2'].alignment = align_center; sheet.merge_cells('J2:K2')
    sheet['L2'].value = data.get('blNo'); sheet['L2'].alignment = align_center; sheet.merge_cells('L2:M2')
    
    bl_type = data.get('blType'); 
    sheet['J3'].value = f"({ 'X' if bl_type == 'EXPRESS RELEASE' else ' ' }) EXPRESS RELEASE ({ 'X' if bl_type == 'TELEX RELEASE' else ' ' }) TELEX RELEASE\n({ 'X' if bl_type == 'WAYBILL' else ' ' }) WAYBILL (NO OBL REQUIRED)"; 
    sheet['J3'].alignment = align_left_center; sheet.merge_cells('J3:M4')

    obl_issued_at = data.get('oblIssuedAt'); 
    sheet['J5'].value = f"OBL ISSUED AT: ({ 'X' if obl_issued_at == 'ORIGIN' else ' ' }) ORIGIN ({ 'X' if obl_issued_at == 'DESTINATION' else ' ' }) DESTINATION"; 
    sheet['J5'].alignment = align_left_center; sheet.merge_cells('J5:M6')
    
    freight_type = data.get('freightType'); 
    sheet['J7'].value = f"FREIGHT: ({ 'X' if freight_type == 'PREPAID' else ' ' }) PREPAID ({ 'X' if freight_type == 'PREPAID ABROAD' else ' ' }) PREPAID ABROAD ({ 'X' if freight_type == 'COLLECT' else ' ' }) COLLECT"; 
    sheet['J7'].alignment = align_left_center; sheet.merge_cells('J7:M8')
    
    # -- Seção 2: Consignee e Notifies --
    style_and_merge('A9:I9', "CONSIGNEE", brand_orange_fill); sheet['A10'].value = data.get('consignee'); sheet['A10'].alignment = align_left_top; sheet.merge_cells('A10:I15')
    style_and_merge('J9:M9', "NOTIFY 2", brand_orange_fill); sheet['J10'].value = data.get('notifyParty2'); sheet['J10'].alignment = align_left_top; sheet.merge_cells('J10:M15')
    style_and_merge('A16:I16', "NOTIFY 1", brand_orange_fill); sheet['A17'].value = data.get('notifyParty'); sheet['A17'].alignment = align_left_top; sheet.merge_cells('A17:I23')

    # -- Seção 3: Logo e Endereço Hevile --
    style_and_merge('J16:M16', "DOMESTIC ROUTING / EXPORT INSTRUCTIONS", brand_blue_fill, font=white_font)

    cell_name = sheet['J17']
    cell_name.value = "HEVILE LOGISTICA E CONSULTORIA INTERNACIONAL LTDA"
    cell_name.font = black_bold_font
    cell_name.alignment = align_center
    sheet.merge_cells('J17:M18')

    cell_address = sheet['J19']
    cell_address.value = "AVENIDA ENGENHEiro DOMINGOS\nFERREIRA, 4661 SL 403/406\nBOA VIAGEM, RECIFE - PE"
    cell_address.font = default_font
    cell_address.alignment = align_center
    sheet.merge_cells('J19:M21')

    sheet.merge_cells('J22:M29')
    try:
        logo_path = os.path.join(BASE_DIR, 'logo.png'); img = Image(logo_path)
        img.height = 112; img.width = 210
        sheet.add_image(img, 'J23')
    except Exception as e:
        print(f"Erro ao adicionar a logo: {e}")
        pass
    
    # -- Demais seções --
    style_and_merge('A24:C24', "PRE-CARRIAGE BY", brand_orange_fill)
    style_and_merge('D24:I24', "PLACE OF RECEIPT", brand_orange_fill)

    sheet.merge_cells('A25:C25')
    sheet.merge_cells('D25:I25')

    style_and_merge('A26:C26', "VESSEL", brand_orange_fill)
    style_and_merge('D26:I26', "PORT OF LOADING", brand_orange_fill)
    style_and_merge('A28:C28', "PORT OF DISCHARGE", brand_orange_fill)
    style_and_merge('D28:I28', "PLACE OF DELIVERY", brand_orange_fill)
    
    sheet['A27'].value = f"{data.get('vessel', '')} / {data.get('voyage', '')}"; sheet['A27'].alignment = align_center; sheet.merge_cells('A27:C27')
    sheet['D27'].value = data.get('portOfLoading'); sheet['D27'].alignment = align_center; sheet.merge_cells('D27:I27')
    sheet['A29'].value = data.get('portOfDischarge'); sheet['A29'].alignment = align_center; sheet.merge_cells('A29:C29')
    sheet['D29'].value = data.get('portOfDischarge'); sheet['D29'].alignment = align_center; sheet.merge_cells('D29:I29')
    style_and_merge('A30:M30', "Wooden Package", brand_orange_fill)
    wood_decl = data.get('woodDeclaration'); sheet['A31'].value = f"({ 'X' if wood_decl == 'Not Applicable' else ' ' }) Not Applicable  ({ 'X' if wood_decl == 'Processed' else ' ' }) Processed  ({ 'X' if wood_decl == 'Treated / Certified' else ' ' }) Treated / Certified  ({ 'X' if wood_decl == 'Not Treated / Not Certified' else ' ' }) Not Treated / Not Certified"; sheet['A31'].alignment = align_center; sheet.merge_cells('A31:M31')
    
    headers_carga = [('A32:B32', "MARKS AND NUMBER"), ('C32:D32', "QUANTITY"), ('E32:H32', "DESCRIPTION OF GOODS (WITH TOTAL NET WEIGHT)"), ('I32:K32', "TOTAL GROSS WEIGHT (KGS)"), ('L32:M32', "TOTAL MEASUREMENT (CBM)")]
    for r, v in headers_carga: style_and_merge(r, v, brand_blue_fill, font=white_font)

    # --- CÁLCULO DOS TOTAIS ---
    total_package_count = 0
    total_gross_weight = 0.0
    total_measurement_cbm = 0.0

    containers_list = data.get('containers', [])
    for container in containers_list:
        try:
            total_package_count += int(container.get('packageCount', 0))
        except (ValueError, TypeError):
            pass
        try:
            total_gross_weight += float(container.get('grossWeight', 0))
        except (ValueError, TypeError):
            pass
        try:
            total_measurement_cbm += float(container.get('measurementCBM', 0))
        except (ValueError, TypeError):
            pass
            
    sheet['A33'].value = data.get('marksAndNumber'); sheet['A33'].alignment = align_left_top; sheet.merge_cells('A33:B47')
    
    sheet['C33'].value = total_package_count
    sheet['C33'].alignment = align_left_top
    sheet.merge_cells('C33:D47')
    
    desc_content = (f"{data.get('cargoDescription', '')}\n\nDUE: {data.get('ducNumber', '')}\nNCM/HS CODE: {data.get('ncmCodes', '')}\nINVOICE: {data.get('exportReferences', '')}\n" + (f"TEMPERATURE: {data.get('temperature', '')}\n" if data.get('temperature') else "") + f"NET WEIGHT: {data.get('netWeight', '')} KGS")
    sheet['E33'].value = desc_content; sheet['E33'].alignment = align_left_top; sheet.merge_cells('E33:H47')
    
    sheet['I33'].value = total_gross_weight
    sheet['I33'].alignment = align_left_top
    sheet.merge_cells('I33:K47')
    
    sheet['L33'].value = total_measurement_cbm
    sheet['L33'].alignment = align_left_top
    sheet.merge_cells('L33:M47')

    # --- SEÇÃO DE MÚLTIPLOS CONTÊINERES ---
    container_headers = [('A48:D48', "CONTAINER Nº"), ('E48:F48', "SEAL"), ('G48', "TARA"), ('H48', "QTY. PACK"), ('I48:J48', "PACK TYPE"), ('K48:L48', "G.WEIGHT"), ('M48', "M/3")]
    for r, v in container_headers: style_and_merge(r, v, brand_blue_fill, font=white_font)

    containers_data = data.get('containers', [])
    start_row = 49

    for i, container in enumerate(containers_data):
        current_row = start_row + i
        
        if current_row >= 70:
            break
            
        sheet.cell(row=current_row, column=1).value = container.get('containerNo')
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        
        sheet.cell(row=current_row, column=5).value = f"ARM: {container.get('sealArmador', '')}\nSHP: {container.get('sealShipper', '')}"
        sheet.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
        
        sheet.cell(row=current_row, column=7).value = container.get('containerTareWeight')
        sheet.cell(row=current_row, column=8).value = container.get('packageCount')
        
        sheet.cell(row=current_row, column=9).value = container.get('packType')
        sheet.merge_cells(start_row=current_row, start_column=9, end_row=current_row, end_column=10)
        
        sheet.cell(row=current_row, column=11).value = container.get('grossWeight')
        sheet.merge_cells(start_row=current_row, start_column=11, end_row=current_row, end_column=12)
        
        sheet.cell(row=current_row, column=13).value = container.get('measurementCBM')

        for col_idx in range(1, 14):
             sheet.cell(row=current_row, column=col_idx).alignment = align_center

    next_empty_row = start_row + len(containers_data)
    for row_idx in range(next_empty_row, 70): 
        for col_idx in range(1, 14):
            sheet.cell(row=row_idx, column=col_idx).border = dotted_border

    # -- Seção 8: Rodapé --
    remarks_text = "RECEIVED BY THE CARRIER FROM THE SHIPPER..." # Texto completo omitido por você
    cell = sheet['A70']; cell.value = remarks_text; cell.alignment = align_left_top; cell.font = small_font; sheet.merge_cells('A70:I79')
    
    cell = sheet['A80']; cell.value = "RECEIPT FOR DELIVERY APPLY TO:"; cell.alignment = align_left_top; sheet.merge_cells('A80:I89')

    place_issue_text = "PLACE AND DATE OF ISSUE"
    date_text = f"Recife, {datetime.now().strftime('%d/%b/%Y')}"
    
    num_to_word = {'01':'ONE', '02':'TWO', '03':'THREE'}
    bl_count_num = data.get('signedBlCount', '')
    bl_word = num_to_word.get(bl_count_num, '').upper()
    signed_text = f"NUMBER OF B(s)/L SIGNED_ {bl_count_num} ({bl_word})"

    full_right_text = ('\n' * 9) + place_issue_text + '\n' + date_text + ('\n' * 7) + signed_text
    
    cell = sheet['J70']
    cell.value = full_right_text
    cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    sheet.merge_cells('J70:M89')

    # --- 4. APLICAR BORDAS (AJUSTADO PARA COLUNAS A-M) ---
    
    def apply_outline_border(cell_range):
        rows = list(sheet[cell_range])
        for cell in rows[0]: cell.border = Border(top=thin_black_border.top, left=thin_black_border.left, right=thin_black_border.right, bottom=thin_black_border.bottom)
        for cell in rows[-1]: cell.border = Border(top=thin_black_border.top, left=thin_black_border.left, right=thin_black_border.right, bottom=thin_black_border.bottom)
        for row in rows:
            row[0].border = Border(top=thin_black_border.top, left=thin_black_border.left, right=thin_black_border.right, bottom=thin_black_border.bottom)
            row[-1].border = Border(top=thin_black_border.top, left=thin_black_border.left, right=thin_black_border.right, bottom=thin_black_border.bottom)

    for row in sheet.iter_rows(min_row=1, max_row=47, min_col=1, max_col=13):
        for cell in row: cell.border = thin_black_border

    for row in sheet.iter_rows(min_row=48, max_row=89, min_col=1, max_col=13):
        for cell in row: cell.border = dotted_border
            
    for cell in sheet["A48":"M48"][0]:
        cell.border = thin_black_border

    apply_outline_border('A70:I79')
    apply_outline_border('A80:I89')
    apply_outline_border('J70:M89')
    
    # --- Salvar e Enviar ---
    virtual_workbook = BytesIO(); workbook.save(virtual_workbook); virtual_workbook.seek(0)
    filename = f"DRAFT HEVILE - {data.get('shipperName', 'documento')}.xlsx"
    return send_file(virtual_workbook, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.sheet")

# --- 4. COMANDO PARA INICIALIZAR O BANCO DE DADOS ---
# Este comando cria as tabelas no banco de dados se elas não existirem
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    # Em produção (no Render), o Render usa um servidor Gunicorn, não este comando.
    # Mas para desenvolvimento local, rodamos assim:
    app.run(host='0.0.0.0', port=5000, debug=True)